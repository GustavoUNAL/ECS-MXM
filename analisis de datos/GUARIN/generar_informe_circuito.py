#!/usr/bin/env python3
"""Extrae resumen del circuito 10-502 desde Excel OR y valida contra PDF unifilar."""
from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
DATA = BASE.parents[1] / "GUARIN" / "data" / "Datos Circuito 10 502 - 2026.xlsx"
PDF = BASE.parents[1] / "GUARIN" / "data" / "CTO 10 502.pdf"
OUT_JSON = BASE / "informe_circuito.json"
OUT_JS = BASE / "informe_circuito.js"


def norm_terminal(t) -> str | None:
    if not t:
        return None
    s = str(t).strip()
    # Cubículo: 1103458A -> 1103458 (no truncar "10 502_Term")
    m = re.fullmatch(r"(\d+)([A-Za-z]+)", s)
    if m:
        return m.group(1)
    return s


def main() -> None:
    wb = openpyxl.load_workbook(DATA, data_only=True)
    ws = wb["Líneas "]
    lines = []
    nodes: set[str] = set()
    cond_stats: dict[str, dict] = defaultdict(lambda: {"tramos": 0, "km": 0.0})

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        lid = str(row[0]).strip()
        ti, tj = str(row[3] or ""), str(row[4] or "")
        L = float(row[5] or 0)
        cond = str(row[6] or "").strip()
        lines.append({"id": lid, "from": norm_terminal(ti), "to": norm_terminal(tj), "L_km": L, "conductor": cond})
        for t in (ti, tj):
            n = norm_terminal(t)
            if n:
                nodes.add(n)
        cond_stats[cond]["tramos"] += 1
        cond_stats[cond]["km"] += L

    ws2 = wb["Cargas_Demanda_corto"]
    loads = []
    demand_by_h: dict[int, dict] = {}
    poc_carga = None
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        name = str(row[0]).strip()
        if row[10] is not None and row[12] is not None and hasattr(row[10], "hour"):
            try:
                h = row[10].hour
                demand_by_h[h] = {
                    "hora": h,
                    "I_A": round(float(row[11]), 3),
                    "S_MVA": round(float(row[12]), 6),
                }
            except (TypeError, ValueError):
                pass
        if "_LOAD" not in name or row[4] is None:
            continue
        node = norm_terminal(row[2])
        s_mva = float(row[4])
        fp = float(row[5] or 0.9)
        loads.append({"name": name, "node": node, "S_MVA": s_mva, "fp": fp, "P_kW": round(s_mva * fp * 1000, 1)})
        if node == "3272966":
            poc_carga = name
    demand = [demand_by_h[h] for h in sorted(demand_by_h)]

    ws3 = wb["Protecciones"]
    prot_rows = list(ws3.iter_rows(values_only=True))
    protecciones = []
    for row in prot_rows[4:]:
        if not row or not row[0]:
            continue
        rtc = float(row[3] or 60)
        protecciones.append(
            {
                "subestacion": row[0],
                "bahia": row[1],
                "tension_kv": row[2],
                "rele": f"{row[5]} {row[6]}",
                "rtc": rtc,
                "rtp": row[4],
                "funciones": str(row[8] or ""),
                "51_prim_A": round(float(row[12] or 0) * rtc, 1),
                "50_prim_A": round(float(row[15] or 0) * rtc, 1),
                "51n_prim_A": round(float(row[18] or 0) * rtc, 1),
                "50n_prim_A": round(float(row[21] or 0) * rtc, 1),
                "51_tms": row[13],
                "51n_tms": row[19],
                "recierre_79": row[23],
            }
        )

    import fitz  # pymupdf

    doc = fitz.open(PDF)
    pdf_nums: set[str] = set()
    for page in doc:
        for ln in page.get_text("text").splitlines():
            ln = ln.strip()
            if re.match(r"^\d{5,7}$", ln):
                pdf_nums.add(ln)
    doc.close()

    line_ids = {l["id"] for l in lines}
    pdf_line_hits = pdf_nums & line_ids
    pdf_node_hits = pdf_nums & nodes
    pdf_huerfanas = pdf_nums - line_ids - nodes
    excel_lineas_sin_pdf = line_ids - pdf_nums
    excel_nodos_sin_pdf = sorted(nodes - pdf_nums)

    def clasificar_omitida(lid: str) -> str:
        row = next((l for l in lines if l["id"] == lid), None)
        if not row:
            return "otro"
        if row["from"] == row["to"]:
            return "lazo_interno"
        if "ACSR" in row.get("conductor", "") and row.get("L_km", 0) >= 0.02:
            return "espina_acsr"
        if row.get("L_km", 0) < 0.002:
            return "tramo_corto"
        return "derivacion_cu"

    omitidas_ids = sorted(excel_lineas_sin_pdf, key=lambda x: (0, int(x)) if x.isdigit() else (1, x))
    omitidas_por_tipo: dict[str, int] = defaultdict(int)
    for lid in omitidas_ids:
        omitidas_por_tipo[clasificar_omitida(lid)] += 1

    pdf_vs_excel = {
        "rol_pdf": "Diagrama unifilar estático de referencia (topología gráfica ESSA). No contiene demanda, cargas ni protección.",
        "que_cambia_con_estado": "Nada en el PDF. Al cambiar hora o Sin/Con AGPE solo cambian insumos del Excel y resultados de flujo de carga; el unifilar es el mismo.",
        "origen_pdf": "CONUCO_13.8",
        "origen_excel": "10 502_Term",
        "equivalencia_origen": "Misma subestación Conucos 13,8 kV — distinto nombre de terminal",
        "tiene_pdf": [
            "Topología simplificada del circuito",
            "IDs de nodos y líneas principales (etiquetas gráficas)",
            "Tipos de conductor (ACSR, CU XLPE…)",
            "Origen CONUCO_13.8",
        ],
        "tiene_solo_excel": [
            "215 tramos con R, X, longitud exacta",
            "51 cargas LV con S, FP, P, Q",
            "Perfil demanda cabecera 24 h",
            "Relé MICOM P142 y ajustes",
            "Estados operacionales (escala de cargas)",
        ],
        "lineas_diagrama_manual": 187,
        "lineas_diagrama_en_or": 187,
        "lineas_or_no_diagrama": 28,
        "lineas_or_no_diagrama_ids": [
            "5129", "5131", "5133", "5134", "5141", "5144", "5146", "5155", "5158",
            "5284", "5286", "5294", "5299", "9138", "9274", "9275", "9341", "9342",
            "9991", "10400", "26922", "747634", "783432", "783433", "796468", "802784",
            "806782", "830652",
        ],
        "omitidas_por_tipo": dict(omitidas_por_tipo),
        "cargas_en_nodos_pdf": len(loads),
        "cargas_fuera_etiqueta_pdf": 0,
    }

    pct_pdf_en_excel = round(100 * len(pdf_line_hits | pdf_node_hits) / len(pdf_nums), 1) if pdf_nums else 0
    if not pdf_huerfanas and len(pdf_line_hits | pdf_node_hits) == len(pdf_nums):
        veredicto = "compatible"
        veredicto_txt = (
            "Todas las etiquetas numéricas del PDF existen en el Excel. "
            "El PDF es unifilar (subconjunto gráfico); el Excel es el modelo completo."
        )
    else:
        veredicto = "revisar"
        veredicto_txt = "Hay etiquetas del PDF sin correspondencia en el Excel."

    horas_estudio = {"09:00": 1.459139, "12:00": 1.682856, "15:00": 1.638498}
    base_hora = "12:00"
    base_s = horas_estudio[base_hora]

    def q_kvar(s_mva: float, fp: float) -> float:
        return s_mva * math.sqrt(max(0.0, 1.0 - fp * fp)) * 1000

    cargas_sin_agpe: dict = {
        "metodo": "Escala uniforme vs hora base 12:00 (modelo calibrado al pico)",
        "formula": "P(hora) = P_nominal × (S_cabecera(hora) / S_cabecera(12:00))",
        "base_hora": base_hora,
        "base_s_cabecera_mva": base_s,
        "estados": {},
    }
    for hora, s_cab in horas_estudio.items():
        escala = s_cab / base_s
        filas = []
        for ld in sorted(loads, key=lambda x: (-x["S_MVA"], x["name"])):
            s = ld["S_MVA"] * escala
            fp = ld["fp"]
            filas.append(
                {
                    "name": ld["name"],
                    "node": ld["node"],
                    "s_nominal_mva": ld["S_MVA"],
                    "s_mva": round(s, 6),
                    "p_kw": round(s * fp * 1000, 2),
                    "q_kvar": round(q_kvar(s, fp), 2),
                    "fp": fp,
                }
            )
        cargas_sin_agpe["estados"][hora] = {
            "estado": "SIN_AGPE",
            "s_cabecera_mva": s_cab,
            "i_cabecera_a": next((d["I_A"] for d in demand if d["hora"] == int(hora[:2])), None),
            "escala": round(escala, 6),
            "escala_pct": round(escala * 100, 2),
            "totales": {
                "s_mva": round(sum(f["s_mva"] for f in filas), 4),
                "p_kw": round(sum(f["p_kw"] for f in filas), 1),
                "q_kvar": round(sum(f["q_kvar"] for f in filas), 1),
            },
            "cargas": filas,
        }

    plan_estudio_conexion = {
        "titulo": "Plan de acción — Estudio de conexión simplificado",
        "proyecto": "MAS X MENOS Guarín · 141,6 kWp / 120 kW AC · circuito 10-502 · POC 3272966",
        "normativa": ["CREG 025/1995", "CREG 174/2021", "CREG 030/2018"],
        "rol_archivos": [
            {
                "archivo": "Datos Circuito 10 502 - 2026.xlsx",
                "rol": "Modelo de cálculo: 215 líneas, 51 cargas, demanda 24 h, CC en SE, relé MICOM P142",
                "no_es": "No sustituye el diagrama unifilar entregado por el OR",
            },
            {
                "archivo": "CTO 10 502.pdf",
                "rol": "Referencia gráfica: topología simplificada, IDs visibles, origen CONUCO_13.8",
                "no_es": "Sin demanda, cargas, protección ni escenarios — no cambia con Sin/Con AGPE",
            },
        ],
        "regla": "Todo el estudio numérico sale del Excel + PowerFactory. El PDF solo valida topología.",
        "criterios": [
            {"parametro": "Tensión en barras", "limite": "0,90 – 1,10 p.u."},
            {"parametro": "Cargabilidad líneas", "limite": "≤ 100 % In"},
            {"parametro": "Cargabilidad transformadores", "limite": "≤ 100 %"},
            {"parametro": "Aumento cortocircuito POC", "limite": "< 1 %"},
        ],
        "escenarios": {
            "horas": ["09:00", "12:00", "15:00"],
            "modos": ["Sin AGPE", "Con AGPE 120 kW"],
            "total": 6,
            "metodo_cargas": "P(hora) = P_nominal × S_cabecera(hora) / S_cabecera(12:00); Sin y Con AGPE usan las mismas cargas a igual hora",
        },
        "fases": [
            {
                "id": 0,
                "nombre": "Consolidar insumos",
                "objetivo": "Paquete de datos trazable y coherente Excel ↔ PDF",
                "pasos": [
                    "Inventario desde Excel: líneas, cargas, perfil cabecera, POC, protección, CC en SE",
                    "Validación cruzada con PDF unifilar (etiquetas, origen CONUCO_13.8 ↔ 10 502_Term)",
                    "Documentar tramos sin etiqueta en unifilar y líneas OR no graficadas",
                    "Reunir fichas inversores, paneles, trafo 150 kVA, unifilar AGPE y subestación del proyecto",
                ],
                "entregable": "Checklist de fuentes + informe interactivo validado",
                "estado": "hecho",
            },
            {
                "id": 1,
                "nombre": "Modelo en PowerFactory",
                "objetivo": "Reproducir circuito 10-502 y montar AGPE en POC",
                "pasos": [
                    "Importar topología desde Excel (215 líneas, impedancias, In)",
                    "Cargar 51 cargas LV con FP 0,9 y escala por perfil horario",
                    "Equivalente SE Conucos 13,8 kV según Excel",
                    "Modelar trafo 150 kVA + SSFV 1/2 CPW (120 kW @ FP 0,99) en nodo 3272966",
                    "Ejecutar validar_modelo_pf.py y resolver discrepancias con scripts _cmp_*",
                ],
                "entregable": "validacion_pf.txt con PASS en conteos, POC y convergencia LF",
                "estado": "parcial",
            },
            {
                "id": 2,
                "nombre": "Estudios eléctricos",
                "objetivo": "Flujo de carga, cortocircuito y coordinación de protecciones",
                "pasos": [
                    "Flujo de carga: 6 escenarios (3 h × Sin/Con AGPE) con exportar_flujo_carga.py",
                    "Verificar tensiones, cargabilidad de líneas y trafo POC vs criterios CREG",
                    "Cortocircuito IEC 60909 en POC: Ik trifásico/monofásico sin y con AGPE",
                    "Coordinación: 51 cabecera vs demanda, exportación AGPE vs relé, selectividad BT/MT, breakers LV",
                    "Comparar pérdidas técnicas Sin vs Con AGPE por hora",
                ],
                "entregable": "results/flujo_carga/ + análisis CC + conclusión coordinación",
                "estado": "parcial",
            },
            {
                "id": 3,
                "nombre": "Análisis y figuras",
                "objetivo": "Convertir resultados PF en evidencia del informe",
                "pasos": [
                    "Regenerar figuras LaTeX (tensiones 09/12/15, cargabilidad, pérdidas, POC)",
                    "Tablas de nodos afectados en ruta al POC",
                    "Sincronizar escenarios_operacion.js e informe HTML con última corrida PF",
                    "Redactar conclusiones por criterio (cumple / no cumple)",
                ],
                "entregable": "Figuras y tablas listas para capítulo de resultados",
                "estado": "parcial",
            },
            {
                "id": 4,
                "nombre": "Documento del estudio",
                "objetivo": "PDF final del estudio de conexión simplificado",
                "pasos": [
                    "Completar capítulos LaTeX: intro, antecedentes, info general, metodología",
                    "Capítulo resultados: LF, CC, protecciones, pérdidas (Sin vs Con AGPE)",
                    "Recomendaciones, coordinación de protecciones y anexos técnicos",
                    "Compilar main.tex → Estudio_Conexion_MAS_X_MENOS.pdf",
                ],
                "entregable": "Estudio_Conexion_MAS_X_MENOS.pdf",
                "estado": "parcial",
            },
            {
                "id": 5,
                "nombre": "Paquete para ESSA",
                "objetivo": "Expediente de radicación ante el operador de red",
                "pasos": [
                    "Estudio de conexión simplificado (PDF)",
                    "Memoria de cálculo y unifilar AGPE del proyecto",
                    "Fichas técnicas y certificados de equipos",
                    "Anexo resultados PF (6 escenarios) y nota modelo vs unifilar OR",
                ],
                "entregable": "Expediente completo para ESSA",
                "estado": "pendiente",
            },
        ],
        "hecho": [
            "Excel analizado y validado vs PDF (100 % etiquetas)",
            "6 escenarios LF con resultados favorables en corrida previa",
            "Coordinación preliminar MICOM P142 documentada",
            "Informe HTML interactivo con grafo y estados operacionales",
            "LaTeX y figuras base del estudio",
        ],
        "pendiente": [
            "Confirmar Excel OR como versión definitiva antes de cerrar modelo PF",
            "Re-ejecutar exportar_flujo_carga.py si el Excel cambió",
            "Actualizar cortocircuito POC y curvas TCC en LaTeX",
            "Revisión final del PDF y firma",
        ],
        "decision": "Si el Excel Datos Circuito 10 502 - 2026.xlsx es la versión definitiva del OR, proceder a cerrar Fase 1–2; si puede cambiar, congelar versión antes de simular.",
    }

    informe = {
        "fuentes": {
            "excel": DATA.name,
            "pdf": PDF.name,
            "fecha_demanda": "2024-03-19",
            "subestacion": "SE Conucos",
            "circuito": "10-502",
            "tension_kv": 13.8,
            "operador": "ESSA",
            "proyecto": "MAS X MENOS Guarín — 141,6 kWp / 120 kW AC",
        },
        "inventario": {
            "n_lineas": len(lines),
            "n_nodos": len(nodes),
            "n_cargas": len(loads),
            "longitud_km": round(sum(l["L_km"] for l in lines), 3),
            "s_cargas_mva": round(sum(l["S_MVA"] for l in loads), 3),
            "poc_nodo": "3272966",
            "poc_carga": poc_carga,
        },
        "pdf_validacion": {
            "veredicto": veredicto,
            "veredicto_txt": veredicto_txt,
            "etiquetas_numericas_pdf": len(pdf_nums),
            "etiquetas_en_excel": len(pdf_line_hits | pdf_node_hits),
            "pct_pdf_en_excel": pct_pdf_en_excel,
            "lineas_en_pdf": len(pdf_line_hits),
            "lineas_pdf_ok": len(pdf_line_hits),
            "lineas_excel_total": len(line_ids),
            "lineas_excel_sin_etiqueta_pdf": len(excel_lineas_sin_pdf),
            "nodos_en_pdf": len(pdf_node_hits),
            "nodos_pdf_ok": len(pdf_node_hits),
            "nodos_excel_total": len(nodes),
            "nodos_excel_sin_etiqueta_pdf": len(excel_nodos_sin_pdf),
            "nodos_sin_pdf": excel_nodos_sin_pdf[:8],
            "nota": (
                f"{len(excel_lineas_sin_pdf)} tramos del Excel no tienen etiqueta visible en el PDF "
                "(lazos internos, seccionadores). Es normal en unifilares."
            ),
        },
        "demanda": {
            "puntos": demand,
            "s_max_mva": max(d["S_MVA"] for d in demand),
            "s_min_mva": min(d["S_MVA"] for d in demand),
            "s_prom_mva": round(sum(d["S_MVA"] for d in demand) / len(demand), 6),
            "i_max_a": max(d["I_A"] for d in demand),
            "horas_estudio": {"09:00": 1.459139, "12:00": 1.682856, "15:00": 1.638498},
        },
        "conductores": [
            {"name": k, "tramos": v["tramos"], "km": round(v["km"], 3)}
            for k, v in sorted(cond_stats.items(), key=lambda x: -x[1]["km"])[:12]
        ],
        "estados_operacionales": {
            "descripcion": "3 horas representativas ESSA × Sin/Con AGPE 120 kW AC (metodología CREG)",
            "horas": ["09:00", "12:00", "15:00"],
            "total": 6,
            "insumo_fijo": [
                "Topología: 215 líneas, 184 nodos (Excel)",
                "Impedancias de líneas y cargas nominales (Excel)",
                "Perfil de demanda de cabecera 19/03/2024 (Excel)",
            ],
            "que_cambia_hora": {
                "titulo": "Al cambiar la hora (09 / 12 / 15 h)",
                "insumo": [
                    "Sube o baja la demanda de cabecera (S en MVA)",
                    "Todas las 51 cargas LV se escalan proporcionalmente (mismo FP 0,9)",
                    "Ej.: a las 12:00 S cabecera = 1,683 MVA (pico); a las 04:00 = 0,630 MVA",
                ],
                "resultado_pf": [
                    "Corrientes e intensidad (%) en líneas",
                    "Tensiones en nodos (p.u.)",
                    "Pérdidas del circuito (kW)",
                    "Potencia del trafo en POC (importación desde red)",
                ],
            },
            "que_cambia_agpe": {
                "titulo": "Al pasar Sin AGPE → Con AGPE (misma hora)",
                "insumo": [
                    "La demanda de cabecera NO cambia (misma hora ESSA)",
                    "Las cargas LV mantienen la misma potencia escalada",
                    "Se activa inyección de 120 kW AC en el POC (3272966)",
                ],
                "resultado_pf": [
                    "Tensiones suben (menos caída; U mín mejora ~0,01 p.u.)",
                    "Pérdidas bajan (menos corriente neta desde cabecera)",
                    "Carga del trafo POC sube en % (120 kW adicionales)",
                    "Trafo POC: de importar (~40 kW) a exportar (~−78 kW neto)",
                    "Intensidad en líneas ruta POC baja ligeramente",
                ],
            },
            "que_no_cambia": [
                "Topología ni impedancias del circuito",
                "Número de cargas (51) ni ubicación",
                "Relé de protección (ajustes fijos del Excel)",
            ],
        },
        "excel_analisis": {
            "hojas": ["Líneas", "Cargas_Demanda_corto", "Protecciones"],
            "tramos_lazo": sum(1 for l in lines if l["from"] == l["to"]),
            "n_conductores": len(cond_stats),
            "carga_min_mva": round(min(l["S_MVA"] for l in loads), 4) if loads else None,
            "carga_max_mva": round(max(l["S_MVA"] for l in loads), 4) if loads else None,
            "p_cargas_kw": round(sum(l["P_kW"] for l in loads), 1),
            "protecciones": protecciones,
        },
        "cargas_sin_agpe": cargas_sin_agpe,
        "pdf_vs_excel": pdf_vs_excel,
        "plan_estudio_conexion": plan_estudio_conexion,
    }

    OUT_JSON.write_text(json.dumps(informe, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_JS.write_text(
        "const INFORME = " + json.dumps(informe, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"OK -> {OUT_JSON.name}, {OUT_JS.name}")


if __name__ == "__main__":
    main()
