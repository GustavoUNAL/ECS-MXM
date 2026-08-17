#!/usr/bin/env python3
"""Análisis completo del Excel OR circuito 10-502."""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA = Path(__file__).resolve().parents[2] / "GUARIN" / "data" / "Datos Circuito 10 502 - 2026.xlsx"
OUT = Path(__file__).resolve().parent / "analisis_excel_10502.json"


def norm_terminal(t) -> str | None:
    if not t:
        return None
    s = str(t).strip()
    m = re.fullmatch(r"(\d+)([A-Za-z]+)", s)
    if m:
        return m.group(1)
    return s


def analizar_lineas(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    headers1, headers2 = rows[0], rows[1]
    data_rows = [r for r in rows[2:] if r[0] is not None]

    nodes: set[str] = set()
    cond_stats: dict[str, dict] = defaultdict(lambda: {"tramos": 0, "km": 0.0})
    in_ka_vals: list[float] = []
    self_loops = 0
    lines = []

    for row in data_rows:
        lid = str(row[0]).strip()
        folder = str(row[1] or "").strip()
        grid = str(row[2] or "").strip()
        ti, tj = str(row[3] or ""), str(row[4] or "")
        L = float(row[5] or 0)
        cond = str(row[6] or "").strip()
        in_ka = float(row[7]) if row[7] is not None else None
        r1, x1 = row[8], row[9]
        r0, x0 = row[10], row[11] if len(row) > 11 else (None, None)

        ni, nj = norm_terminal(ti), norm_terminal(tj)
        if ni:
            nodes.add(ni)
        if nj:
            nodes.add(nj)
        if ni and nj and ni == nj:
            self_loops += 1

        cond_stats[cond]["tramos"] += 1
        cond_stats[cond]["km"] += L
        if in_ka is not None:
            in_ka_vals.append(in_ka)

        lines.append(
            {
                "id": lid,
                "folder": folder,
                "grid": grid,
                "from": ni,
                "to": nj,
                "ti_raw": ti,
                "tj_raw": tj,
                "km": L,
                "conductor": cond,
                "in_ka": in_ka,
            }
        )

    by_folder = Counter(l["folder"] or "(sin carpeta)" for l in lines)
    conductores = sorted(
        [{"name": k, "tramos": v["tramos"], "km": round(v["km"], 4)}
         for k, v in cond_stats.items()],
        key=lambda x: -x["km"],
    )

    return {
        "hoja": "Líneas",
        "encabezados_fila1": [str(x) if x is not None else "" for x in headers1],
        "encabezados_fila2": [str(x) if x is not None else "" for x in headers2],
        "n_lineas": len(lines),
        "n_nodos": len(nodes),
        "longitud_km": round(sum(l["km"] for l in lines), 4),
        "tramos_lazo_mismo_nodo": self_loops,
        "n_conductores_distintos": len(cond_stats),
        "conductores": conductores,
        "carpetas_distintas": len(by_folder),
        "por_carpeta": dict(by_folder.most_common(15)),
        "in_ka_min": min(in_ka_vals) if in_ka_vals else None,
        "in_ka_max": max(in_ka_vals) if in_ka_vals else None,
    }


def analizar_cargas(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    loads = []
    demand_by_h: dict[int, dict] = {}
    for row in rows[2:]:
        if row[10] is not None and row[12] is not None:
            try:
                t = row[10]
                if not hasattr(t, "hour"):
                    continue
                h = t.hour
                demand_by_h[h] = {
                    "hora": h,
                    "I_A": round(float(row[11]), 3),
                    "S_MVA": round(float(row[12]), 6),
                }
            except (TypeError, ValueError):
                pass
    demand = [demand_by_h[h] for h in sorted(demand_by_h)]

    for row in rows[2:]:
        if row[0] is None:
            continue
        name = str(row[0]).strip()
        if row[4] is not None and "_LOAD" in name:
            try:
                s_mva = float(row[4])
            except (TypeError, ValueError):
                continue
            node = norm_terminal(row[2])
            fp = float(row[5]) if row[5] is not None else 0.9
            loads.append(
                {
                    "name": name,
                    "bus": str(row[3] or "").strip(),
                    "node": node,
                    "ti_raw": str(row[2] or "").strip(),
                    "S_MVA": round(s_mva, 6),
                    "fp": fp,
                    "P_kW": round(s_mva * fp * 1000, 2),
                }
            )
    s_cargas = sum(l["S_MVA"] for l in loads)
    p_cargas = sum(l["P_kW"] for l in loads)
    poc = next((l for l in loads if l["node"] == "3272966"), None)

    by_fp = Counter(l["fp"] for l in loads if l["fp"] is not None)
    sizes = sorted(l["S_MVA"] for l in loads)

    return {
        "hoja": "Cargas_Demanda_corto",
        "n_cargas": len(loads),
        "s_cargas_mva": round(s_cargas, 4),
        "p_cargas_kw": round(p_cargas, 2),
        "fp_distintos": dict(by_fp),
        "carga_min_mva": min(sizes) if sizes else None,
        "carga_max_mva": max(sizes) if sizes else None,
        "poc": poc,
        "cargas_mayores_0_3_mva": sum(1 for l in loads if l["S_MVA"] >= 0.3),
        "demanda_cabecera": {
            "fecha": "2024-03-19",
            "puntos_24h": len(demand),
            "s_max_mva": max(d["S_MVA"] for d in demand) if demand else None,
            "s_min_mva": min(d["S_MVA"] for d in demand) if demand else None,
            "s_prom_mva": round(sum(d["S_MVA"] for d in demand) / len(demand), 6) if demand else None,
            "i_max_a": max(d["I_A"] for d in demand) if demand else None,
            "horas_estudio": {
                "09:00": next((d["S_MVA"] for d in demand if d["hora"] == 9), None),
                "12:00": next((d["S_MVA"] for d in demand if d["hora"] == 12), None),
                "15:00": next((d["S_MVA"] for d in demand if d["hora"] == 15), None),
            },
        },
        "lista_cargas": loads,
    }


def analizar_protecciones(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    # Fila 3 = encabezados principales, fila 4 = sub-encabezados ajustes
    h_main = [str(x).replace("\n", " ") if x is not None else "" for x in rows[2]]
    h_sub = [str(x).replace("\n", " ") if x is not None else "" for x in rows[3]]

    elementos = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        el = {
            "subestacion": row[0],
            "bahia": row[1],
            "tension_kv": row[2],
            "rtc": row[3],
            "rtp": row[4],
            "fabricante": row[5],
            "modelo": row[6],
            "serial": row[7],
            "protecciones_activas": row[8],
            "funcion": row[9],
            "tipo": row[10],
            "ajustes_fase": {
                "dir": row[11],
                "51_sec": row[12],
                "51_dial_tms": row[13],
                "51_curva": row[14],
                "50_sec": row[15],
                "50_delay_s": row[16],
            },
            "ajustes_tierra": {
                "dir": row[17],
                "51n_sec": row[18],
                "51n_dial_tms": row[19],
                "51n_curva": row[20],
                "50n_sec": row[21],
                "50n_delay_s": row[22],
            },
            "recierre": {
                "79": row[23],
                "reset_s": row[24],
            },
        }
        # Valores primarios (RTC/RTP)
        rtc, rtp = float(row[3] or 60), float(row[4] or 120)
        el["primarios"] = {
            "51_A": round(float(row[12] or 0) * rtc, 1),
            "50_A": round(float(row[15] or 0) * rtc, 1),
            "51N_A": round(float(row[18] or 0) * rtc, 1),
            "50N_A": round(float(row[21] or 0) * rtc, 1),
        }
        funciones = []
        prot_str = str(row[8] or "")
        for tag in ["51", "50", "51N", "50N", "79"]:
            if tag.lower() in prot_str.lower() or tag in prot_str:
                funciones.append(tag)
        el["funciones_detectadas"] = funciones
        elementos.append(el)

    return {
        "hoja": "Protecciones",
        "encabezados": h_main,
        "n_reles": len(elementos),
        "n_funciones_totales": sum(len(e["funciones_detectadas"]) for e in elementos),
        "elementos": elementos,
    }


def main() -> None:
    wb = openpyxl.load_workbook(DATA, data_only=True)
    lineas = analizar_lineas(wb["Líneas "])
    cargas = analizar_cargas(wb["Cargas_Demanda_corto"])
    protecciones = analizar_protecciones(wb["Protecciones"])

    informe = {
        "archivo": DATA.name,
        "circuito": "10-502",
        "subestacion": "SE Conucos",
        "resumen": {
            "lineas": lineas["n_lineas"],
            "nodos": lineas["n_nodos"],
            "longitud_km": lineas["longitud_km"],
            "cargas": cargas["n_cargas"],
            "s_cargas_mva": cargas["s_cargas_mva"],
            "protecciones": protecciones["n_reles"],
            "funciones_proteccion": protecciones["n_funciones_totales"],
        },
        "lineas": lineas,
        "cargas": {k: v for k, v in cargas.items() if k != "lista_cargas"},
        "lista_cargas": cargas["lista_cargas"],
        "protecciones": protecciones,
    }

    OUT.write_text(json.dumps(informe, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(informe["resumen"], ensure_ascii=False, indent=2))
    print("\n--- PROTECCIONES ---")
    for el in protecciones["elementos"]:
        print(json.dumps(el, ensure_ascii=False))
    print(f"\nGuardado: {OUT.name}")


if __name__ == "__main__":
    main()
