#!/usr/bin/env python3
"""Genera figuras y resumen JSON a partir de datos del circuito 10-502."""
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "GUARIN/data/Datos Circuito 10 502 - 2026.xlsx"
FIG = ROOT / "proyecto/figuras"
OUT = ROOT / "proyecto/analisis"
FIG.mkdir(exist_ok=True)
OUT.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(DATA, data_only=True)

ws = wb["Líneas "]
lines = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    lines.append(
        {
            "name": str(row[0]).strip(),
            "ti": str(row[3]).strip() if row[3] is not None else "",
            "tj": str(row[4]).strip() if row[4] is not None else "",
            "L_km": float(row[5] or 0),
            "cond": str(row[6] or "").strip(),
            "In_kA": float(row[7] or 0),
            "R1": float(row[8] or 0),
            "X1": float(row[9] or 0),
            "B1": float(row[10] or 0),
            "R0": float(row[11] or 0),
            "X0": float(row[12] or 0),
            "B0": float(row[13] or 0),
        }
    )

ws2 = wb["Cargas_Demanda_corto"]
loads = []
for row in ws2.iter_rows(min_row=3, values_only=True):
    if row[0] is None or row[4] is None:
        continue
    loads.append(
        {
            "name": str(row[0]).strip(),
            "terminal": str(row[2]).strip() if row[2] is not None else "",
            "S_MVA": float(row[4]),
            "fp": float(row[5] or 0.9),
        }
    )

demand = []
for row in ws2.iter_rows(min_row=2, max_row=25, values_only=True):
    if row[10] is None or row[12] is None:
        continue
    h = row[10].hour if hasattr(row[10], "hour") else int(str(row[10])[11:13])
    demand.append({"hora": h, "I_A": float(row[11]), "S_MVA": float(row[12])})

path_ids = ["807671", "807672", "807673", "807674", "807675"]
order = {n: i for i, n in enumerate(path_ids)}
path = sorted([l for l in lines if l["name"] in path_ids], key=lambda x: order[x["name"]])
Zpath_R = sum(p["R1"] for p in path)
Zpath_X = sum(p["X1"] for p in path)
Zpath_R0 = sum(p["R0"] for p in path)
Zpath_X0 = sum(p["X0"] for p in path)
Lpath = sum(p["L_km"] for p in path)

Sk3, Ik3, Ip3 = 654.488, 27.382, 70.401
Sk1, Ik1, Ip1 = 239.681, 30.083, 77.345
Vll = 13.8
Zth = (Vll**2) / Sk3
rx = 0.068
Xth = Zth / math.sqrt(1 + rx**2)
Rth = rx * Xth
Rp, Xp = Rth + Zpath_R, Xth + Zpath_X
Zp = math.sqrt(Rp**2 + Xp**2)
Ik3_poc_calc = (Vll / math.sqrt(3)) / Zp

I_nom_lv_A = 120 / (math.sqrt(3) * 0.22)
I_sc_inv_A = 1.5 * I_nom_lv_A
I_sc_inv_hv_A = I_sc_inv_A * (0.22 / 13.2)

S_cli = 0.15
fp = 0.9
P_cli = S_cli * fp * 1000
P_gen = 120
hours = list(range(24))
pv_raw = [math.sin((h - 6) / 12 * math.pi) if 6 <= h <= 18 else 0.0 for h in hours]
mx = max(pv_raw) or 1
pv = [p * P_gen / mx for p in pv_raw]

S_peak = max(d["S_MVA"] for d in demand)
S_model = sum(l["S_MVA"] for l in loads)

cond_count = Counter(l["cond"] for l in lines)
cond_len = defaultdict(float)
cond_In = {}
for l in lines:
    cond_len[l["cond"]] += l["L_km"]
    cond_In[l["cond"]] = l["In_kA"]

top = sorted(loads, key=lambda x: -x["S_MVA"])[:12]
summary = {
    "n_lineas": len(lines),
    "longitud_km": round(sum(l["L_km"] for l in lines), 3),
    "n_cargas": len(loads),
    "S_cargas_MVA": round(S_model, 3),
    "S_max_cabecera_MVA": round(S_peak, 3),
    "I_max_A": round(max(d["I_A"] for d in demand), 2),
    "S_min_MVA": round(min(d["S_MVA"] for d in demand), 3),
    "POC": "3272966",
    "linea_conexion": path[-1],
    "ruta_poc": path,
    "Zpath_ohm": {
        "R1": round(Zpath_R, 4),
        "X1": round(Zpath_X, 4),
        "R0": round(Zpath_R0, 4),
        "X0": round(Zpath_X0, 4),
        "L_km": round(Lpath, 4),
    },
    "Ik3_poc_approx_kA": round(Ik3_poc_calc, 3),
    "I_nom_lv_A": round(I_nom_lv_A, 2),
    "I_sc_inv_lv_A": round(I_sc_inv_A, 2),
    "I_sc_inv_hv_A": round(I_sc_inv_hv_A, 3),
    "P_cli_kW": P_cli,
    "conductores": {
        k: {"tramos": v, "km": round(cond_len[k], 3), "In_kA": cond_In[k]}
        for k, v in cond_count.most_common()
    },
    "demanda": demand,
    "pv_profile": [{"hora": h, "P_kW": round(pv[h], 2)} for h in hours],
    "sc_se": {
        "3ph": {"Sk": Sk3, "Ik": Ik3, "Ip": Ip3},
        "1ph": {"Sk": Sk1, "Ik": Ik1, "Ip": Ip1},
    },
    "top_cargas": top,
}
(OUT / "summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
print(json.dumps({k: summary[k] for k in list(summary)[:12]}, indent=2, default=str))

plt.rcParams.update(
    {
        "font.family": "DejaVu Sans",
        "axes.spines.top": False,
        "axes.spines.right": False,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "axes.grid": True,
        "grid.alpha": 0.25,
    }
)

# Demanda
fig, ax1 = plt.subplots(figsize=(9, 4.2))
hs = [d["hora"] for d in demand]
Ss = [d["S_MVA"] for d in demand]
Is = [d["I_A"] for d in demand]
ax1.fill_between(hs, Ss, alpha=0.25, color="#1f4e79")
ax1.plot(hs, Ss, "o-", color="#1f4e79", lw=2, ms=4, label="S cabecera [MVA]")
ax2 = ax1.twinx()
ax2.spines["top"].set_visible(False)
ax2.plot(hs, Is, "s--", color="#c45c26", lw=1.6, ms=3.5, label="I cabecera [A]")
for hmark, lab in [(9, "9:00"), (12, "12:00"), (15, "15:00")]:
    ax1.axvline(hmark, color="#666", ls=":", lw=1)
    ax1.text(hmark + 0.1, max(Ss) * 0.97, lab, fontsize=8, color="#444")
ax1.set_xlabel("Hora del día")
ax1.set_ylabel("Potencia aparente [MVA]", color="#1f4e79")
ax2.set_ylabel("Corriente [A]", color="#c45c26")
ax1.set_title("Perfil de demanda del circuito 10-502 — SE Conucos (19/03/2024)")
ax1.set_xticks(range(0, 24, 2))
l1, lb1 = ax1.get_legend_handles_labels()
l2, lb2 = ax2.get_legend_handles_labels()
ax1.legend(l1 + l2, lb1 + lb2, loc="upper left", frameon=False)
fig.tight_layout()
fig.savefig(FIG / "fig_demanda_circuito.png", dpi=160)
plt.close()

# Gen vs demanda
fig, ax = plt.subplots(figsize=(9, 4.2))
ax.fill_between(hours, pv, alpha=0.3, color="#2a9d8f", label="Generación FV estimada [kW]")
ax.plot(hours, pv, color="#2a9d8f", lw=2)
ax.axhline(P_cli, color="#e76f51", ls="--", lw=1.8, label=f"Demanda cliente POC ({P_cli:.0f} kW)")
ax.axhline(P_gen, color="#264653", ls=":", lw=1.4, label=f"Pnom AC ({P_gen:.0f} kW)")
export = [h for h in hours if pv[h] > P_cli]
if export:
    ax.axvspan(min(export), max(export), color="#2a9d8f", alpha=0.08)
    ax.text(
        np.mean(export),
        P_gen * 0.55,
        "Ventana de\nexportación",
        ha="center",
        fontsize=9,
        color="#1b6b5e",
    )
ax.set_xlabel("Hora del día")
ax.set_ylabel("Potencia [kW]")
ax.set_title("Generación fotovoltaica estimada vs demanda del cliente (nodo 3272966)")
ax.set_xticks(range(0, 24, 2))
ax.legend(frameon=False, loc="upper left")
fig.tight_layout()
fig.savefig(FIG / "fig_gen_vs_demanda.png", dpi=160)
plt.close()

# Top cargas
fig, ax = plt.subplots(figsize=(9, 4.5))
names = [t["name"].replace("_LV_LOAD", "") for t in top][::-1]
vals = [t["S_MVA"] * 1000 for t in top][::-1]
colors = ["#c45c26" if t["terminal"] == "3272966" else "#1f4e79" for t in top][::-1]
ax.barh(names, vals, color=colors)
ax.set_xlabel("Potencia aparente [kVA]")
ax.set_title("Principales cargas modeladas en el circuito 10-502")
for y, v in enumerate(vals):
    ax.text(v + 5, y, f"{v:.0f}", va="center", fontsize=8)
fig.tight_layout()
fig.savefig(FIG / "fig_top_cargas.png", dpi=160)
plt.close()

# Conductores
fig, ax = plt.subplots(figsize=(9, 4.2))
items = sorted(cond_len.items(), key=lambda x: -x[1])
labs = [k.replace("3F_15_", "") for k, _ in items]
kms = [v for _, v in items]
ax.bar(range(len(labs)), kms, color="#1f4e79")
ax.set_xticks(range(len(labs)))
ax.set_xticklabels(labs, rotation=35, ha="right", fontsize=8)
ax.set_ylabel("Longitud acumulada [km]")
ax.set_title("Composición de conductores del circuito 10-502")
for i, v in enumerate(kms):
    ax.text(i, v + 0.02, f"{v:.2f}", ha="center", fontsize=7)
fig.tight_layout()
fig.savefig(FIG / "fig_conductores.png", dpi=160)
plt.close()

w = 0.35
nodos = ["3272869", "3272966", "3272761", "3272664", "3272567", "1065971", "2479567"]
sin12 = [0.99432, 0.99432, 0.99432, 0.99433, 0.99433, 0.99433, 0.99433]
con12 = [0.99494, 0.99494, 0.99494, 0.99494, 0.99494, 0.99493, 0.99493]
fig, ax = plt.subplots(figsize=(9, 4.2))
x = np.arange(len(nodos))
ax.bar(x - w / 2, sin12, w, label="Sin SSFV", color="#8d99ae")
ax.bar(x + w / 2, con12, w, label="Con SSFV", color="#1f4e79")
ax.axhline(0.90, color="#e76f51", ls="--", lw=1, label="Límite 0.90 p.u.")
ax.axhline(1.10, color="#e76f51", ls=":", lw=1, label="Límite 1.10 p.u.")
ax.set_ylim(0.88, 1.12)
ax.set_xticks(x)
ax.set_xticklabels(nodos, rotation=30, ha="right", fontsize=8)
ax.set_ylabel("Tensión [p.u.]")
ax.set_title("Perfil de tensiones a las 12:00 — nodos cercanos al POC")
ax.legend(frameon=False, fontsize=8, ncol=2)
fig.tight_layout()
fig.savefig(FIG / "fig_tensiones_12.png", dpi=160)
plt.close()

lineas = ["5291", "5292", "5294", "5295", "5296", "26921", "5373"]
sin_c = [4.565, 4.565, 4.565, 4.566, 4.566, 2.812, 4.566]
con_c = [2.839, 2.840, 2.840, 2.841, 2.841, 1.303, 2.842]
fig, ax = plt.subplots(figsize=(9, 4.2))
x = np.arange(len(lineas))
ax.bar(x - w / 2, sin_c, w, label="Sin SSFV", color="#8d99ae")
ax.bar(x + w / 2, con_c, w, label="Con SSFV", color="#2a9d8f")
ax.set_xticks(x)
ax.set_xticklabels(lineas)
ax.set_ylabel("Cargabilidad [%]")
ax.set_title("Cargabilidad de líneas a las 12:00 (máxima variación)")
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(FIG / "fig_cargabilidad_lineas.png", dpi=160)
plt.close()

horas_lab = ["9:00", "12:00", "15:00"]
x = np.arange(3)
sin_t = [29.645, 28.598, 27.000]
con_t = [53.655, 54.403, 55.833]
fig, ax = plt.subplots(figsize=(7.5, 4.2))
ax.plot(x, sin_t, "o-", color="#8d99ae", lw=2, label="Sin SSFV")
ax.plot(x, con_t, "s-", color="#1f4e79", lw=2, label="Con SSFV")
ax.axhline(100, color="#e76f51", ls="--", label="Límite 100%")
ax.fill_between(x, sin_t, con_t, alpha=0.15, color="#1f4e79")
ax.set_xticks(x)
ax.set_xticklabels(horas_lab)
ax.set_ylabel("Cargabilidad [%]")
ax.set_title("Cargabilidad del transformador de conexión (150 kVA)")
ax.set_ylim(0, 110)
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(FIG / "fig_cargabilidad_trafo.png", dpi=160)
plt.close()

sin_p = [4.77, 5.37, 6.03]
con_p = [4.47, 5.03, 5.69]
fig, ax = plt.subplots(figsize=(7.5, 4.2))
ax.bar(x - w / 2, sin_p, w, label="Sin SSFV", color="#8d99ae")
ax.bar(x + w / 2, con_p, w, label="Con SSFV", color="#2a9d8f")
ax.set_xticks(x)
ax.set_xticklabels(horas_lab)
ax.set_ylabel("Pérdidas [kW]")
ax.set_title("Pérdidas técnicas del sistema — comparación con/sin SSFV")
for i, (a, b) in enumerate(zip(sin_p, con_p)):
    ax.annotate(f"-{a - b:.2f} kW", (i, max(a, b) + 0.15), ha="center", fontsize=8, color="#1b6b5e")
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(FIG / "fig_perdidas.png", dpi=160)
plt.close()

nodos_cc = ["3272869", "3272966", "3272761", "3272664", "3272567"]
cc3 = [5.12685, 5.12027, 5.13564, 5.15265, 5.19351]
cc1 = [3.15235, 3.14984, 3.15571, 3.16218, 3.17766]
fig, ax = plt.subplots(figsize=(9, 4.2))
x = np.arange(len(nodos_cc))
ax.bar(x - w / 2, cc3, w, label="Trifásico Ik", color="#1f4e79")
ax.bar(x + w / 2, cc1, w, label="Monofásico Ik", color="#c45c26")
ax.set_xticks(x)
ax.set_xticklabels(nodos_cc, rotation=20, ha="right")
ax.set_ylabel("Corriente de falla [kA]")
ax.set_title("Niveles de cortocircuito en nodos cercanos al POC (12:00)")
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(FIG / "fig_cortocircuito.png", dpi=160)
plt.close()

fig, ax = plt.subplots(figsize=(10, 3.2))
ax.set_xlim(0, 10)
ax.set_ylim(0, 3)
ax.axis("off")
ax.set_title("Esquema unifilar simplificado del punto de conexión", pad=10)
boxes = [
    (0.3, 1.2, 1.6, 1.0, "SE Conucos\n13,8 kV\nCTO 10-502"),
    (2.4, 1.2, 1.4, 1.0, "Tramo MT\nACSR / XLPE"),
    (4.3, 1.2, 1.5, 1.0, "Nodo\n3272869"),
    (6.3, 1.2, 1.6, 1.0, "POC\n3272966\n3 m CU#2"),
    (8.3, 1.2, 1.5, 1.0, "Trafo 150 kVA\n13,2/0,22 kV\n+ SSFV 120 kW"),
]
for x0, y0, w0, h0, txt in boxes:
    ax.add_patch(
        plt.Rectangle((x0, y0), w0, h0, fill=True, facecolor="#e8eef5", edgecolor="#1f4e79", lw=1.5)
    )
    ax.text(x0 + w0 / 2, y0 + h0 / 2, txt, ha="center", va="center", fontsize=8, color="#1f4e79")
for i in range(len(boxes) - 1):
    x1 = boxes[i][0] + boxes[i][2]
    x2 = boxes[i + 1][0]
    ax.annotate(
        "",
        xy=(x2, 1.7),
        xytext=(x1, 1.7),
        arrowprops=dict(arrowstyle="->", color="#c45c26", lw=1.8),
    )
ax.text(
    5,
    0.45,
    "Carga cliente modelada: 150 kVA @ FP 0,9  |  Generación AC: 120 kW (2×60 kW SOLIS)",
    ha="center",
    fontsize=8,
    color="#333",
)
fig.tight_layout()
fig.savefig(FIG / "fig_esquema_poc.png", dpi=160)
plt.close()

print("OK figs:", sorted(p.name for p in FIG.glob("fig_*.png")))
